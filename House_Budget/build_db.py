"""
整合三份 Excel 為 SQLite House Budget 資料庫
"""
import io
import sqlite3
import openpyxl
import os
import sys
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE, "house_budget.db")

FILES = {
    "AU Travel Budget.xlsx":        "au_travel",
    "Sam Household Budget_2025.xlsx": "household_2025",
    "Sam Household Budget_2026.xlsx": "household_2026",
}

# ── helpers ──────────────────────────────────────────────────────────────────

def safe_str(v):
    if v is None:
        return None
    try:
        return str(v).strip() or None
    except Exception:
        return None

def normalize_who(v):
    who = safe_str(v)
    if who is None or who == "阿Sam":
        return "Sam"
    if who in ("cynical Tsai", "cynical07@gmail.com"):
        return "Rita"
    return who

def safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("="):   # Excel formula not evaluated
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None

def safe_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    try:
        return str(v)[:10]
    except Exception:
        return None

# ── schema ───────────────────────────────────────────────────────────────────

CREATE_TRANSACTIONS = """
CREATE TABLE IF NOT EXISTS transactions (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    source      TEXT NOT NULL,   -- file + sheet
    date        TEXT,
    category    TEXT,
    item        TEXT,
    amount      REAL,
    currency    TEXT,
    amount_twd  REAL,            -- TWD equivalent if available
    who         TEXT,             -- payer
    note        TEXT
);
"""

CREATE_SUMMARY = """
CREATE TABLE IF NOT EXISTS monthly_summary AS
SELECT
    source,
    substr(date, 1, 7) AS month,
    category,
    currency,
    COUNT(*)            AS tx_count,
    SUM(amount)         AS total_amount,
    SUM(amount_twd)     AS total_twd
FROM transactions
WHERE date IS NOT NULL
GROUP BY source, month, category, currency;
"""

# ── readers ──────────────────────────────────────────────────────────────────

def read_sheet_rows(ws):
    """Return list of dicts from a sheet.
    Handles two layouts:
      A) Date|Category|Item|Amount|Currency|Who                (most sheets)
      B) Date|Category|Item|Amount|Currency|TWD|Who|Note/Where (Apr 2025+)
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row (first non-empty row)
    header = None
    data_start = 0
    for i, row in enumerate(rows):
        clean = [safe_str(c) for c in row if safe_str(c)]
        if clean and clean[0].lower() in ("date",):
            header = [safe_str(c) for c in row]
            data_start = i + 1
            break

    if header is None:
        return []

    # Detect column positions
    h_lower = [h.lower() if h else "" for h in header]

    def col(name):
        for i, h in enumerate(h_lower):
            if h == name:
                return i
        return None

    idx_date     = col("date")
    idx_category = col("category")
    idx_item     = col("item")
    idx_amount   = col("amount")
    idx_currency = col("currency")
    idx_who      = col("who")
    # TWD column only in some sheets (header "twd")
    idx_twd = col("twd")
    # Note / Where / Column 12 (location tag like "Sendai")
    idx_note = None
    for i, h in enumerate(h_lower):
        if h in ("column 12", "note", "where"):
            idx_note = i
            break

    records = []
    for row in rows[data_start:]:
        if not any(v is not None for v in row):
            continue
        date     = safe_date(row[idx_date])     if idx_date is not None     else None
        category = safe_str(row[idx_category])  if idx_category is not None else None
        item     = safe_str(row[idx_item])       if idx_item is not None     else None
        amount   = safe_float(row[idx_amount])   if idx_amount is not None   else None
        currency = safe_str(row[idx_currency])   if idx_currency is not None else None
        who      = normalize_who(row[idx_who])   if idx_who is not None      else "Sam"
        note     = safe_str(row[idx_note])       if idx_note is not None     else None

        # TWD equivalent
        amount_twd = None
        if idx_twd is not None:
            amount_twd = safe_float(row[idx_twd])
        elif currency in (None, "TWD"):
            amount_twd = amount

        # Skip rows that are clearly formula summaries or empty data
        if date is None and category is None and item is None:
            continue

        records.append({
            "date":       date,
            "category":   category,
            "item":       item,
            "amount":     amount,
            "currency":   currency,
            "amount_twd": amount_twd,
            "who":        who,
            "note":       note,
        })
    return records

# ── main ─────────────────────────────────────────────────────────────────────

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    print("[ 1/4 ] 建立 SQLite 資料庫 …")
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(CREATE_TRANSACTIONS)

    total = 0
    print("[ 2/4 ] 讀取 Excel 檔案 …")
    for filename, source_prefix in FILES.items():
        fpath = os.path.join(BASE, filename)
        if not os.path.exists(fpath):
            print(f"  WARN  file not found: {filename}, skipping")
            continue
        wb = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            source = f"{source_prefix}/{sheet_name.strip()}"
            records = read_sheet_rows(ws)
            for r in records:
                cur.execute("""
                    INSERT INTO transactions
                        (source, date, category, item, amount, currency, amount_twd, who, note)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (
                    source, r["date"], r["category"], r["item"],
                    r["amount"], r["currency"], r["amount_twd"],
                    r["who"], r["note"],
                ))
            total += len(records)
            print(f"  OK  {filename} / {sheet_name}: {len(records)} rows")
        wb.close()

    print(f"\n[ 3/4 ] 建立 monthly_summary table …")
    cur.execute(CREATE_SUMMARY)

    conn.commit()
    conn.close()
    print(f"[ 4/4 ] 完成！共匯入 {total} 筆交易 → {DB_PATH}")
    return total

if __name__ == "__main__":
    main()
